from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
import structlog
from rapidfuzz import process, fuzz

from metrics import quality, hash_file
from utils import parse_price, normalize_unit, is_junk_row, read_csv_iter, simple_match_header, collect_prices_for_row, \
    pick_best_price

logger = structlog.get_logger()


@dataclass
class ParseResult:
    file_path: str
    file_hash: str
    items: list[dict] = field(default_factory=list)
    confidence: float = 0.0
    needs_ai: bool = False
    errors: list[str] = field(default_factory=list)
    method: str = "none"
    col_map: dict = field(default_factory=dict)


class ExcelParser:
    """
    XLSX/XLS/CSV парсер:
    - основной метод: pandas + fuzzy по шапке
    - fallback: простой эвристический парсер по строкам
    - метрики качества _quality + needs_ai
    - PDF/DOC/DOCX нет
    """

    # публичный метод

    def parse_file(self, file_path: str) -> ParseResult:
        """Метод фактического парсинга файла, возвращает объект ParseResult с данными о файле и содержимом."""

        f_path = Path(file_path)
        result = ParseResult(file_path=file_path, file_hash=hash_file(f_path))

        # пробуем парсить файл в зависимости от его формата
        try:

            file_type = f_path.suffix.lower()

            if file_type in (".xlsx", ".xlsm", ".xls"):
                self._parse_spreadsheet(f_path, result, [pd.ExcelFile])
            elif file_type in ['.csv', '.tsv']:
                self._parse_spreadsheet(f_path, result, [], self._standard_csv_fallback)
            else:
                result.errors.append(f"Unsupported type: {file_type}.")
                return result

            AI_FALLBACK_THRESHOLD = 0.55
            result.needs_ai = (result.confidence < AI_FALLBACK_THRESHOLD or not result.items)

        except Exception as e:
            logger.error("parse_error", file=str(f_path), error=str(e))
            result.errors.append(str(e))
            result.needs_ai = True

        return result

    # основной метод парсинга

    def _parse_spreadsheet(self, file_path: Path, result: ParseResult, readers: list, fallback_method: callable = None):
        """Универсальный парсер для spreadsheet-форматов (xlsx, xls, csv)."""

        file_type: str = file_path.suffix.lower()
        col_map: dict = {}

        try:
            if file_type == ".csv":

                # CSV требует попыток разных разделителей
                for sep in [";", ",", "\t", "|"]:

                    try:
                        df_iterator = read_csv_iter(str(file_path), sep)
                        sheet_items: list[dict] = []

                        for df in df_iterator:
                            # если колонок больше, чем 2, парсим с fuzzy
                            if df.shape[1] < 2:
                                continue

                            items, col_map = self._parse_dataframe_with_fuzzy(df)
                            if items:
                                sheet_items.extend(items)

                        if sheet_items:
                            self._handle_success(result, sheet_items, col_map, f"{file_type}+fuzzy", sep)
                            logger.info(f"{file_type}_fuzzy_ok", items=len(sheet_items),
                                        conf=f"{result.confidence:.2f}",
                                        sep=sep)
                            return

                    except Exception as e:
                        result.errors.append(f"{file_type}({sep}): {e}")
                        continue
            else:
                # excel-форматы: один reader, multi-sheet

                if file_type == ".xlsx":
                    fallback_method = self._standard_xlsx_fallback
                elif file_type == ".xls":
                    fallback_method = self._standard_xls_fallback

                xls = readers[0](str(file_path))  # pd.ExcelFile для xlsx/xls
                sheet_items = []

                # проходимся по каждому листу
                for sheet_name in xls.sheet_names:
                    df = xls.parse(sheet_name=sheet_name, dtype=str, header=None)
                    items, col_map = self._parse_dataframe_with_fuzzy(df)

                    if items:
                        sheet_items.extend(items)
                        if not result.col_map:
                            result.col_map = col_map

                if sheet_items:
                    self._handle_success(result, sheet_items, col_map, "fuzzy")
                    logger.info(f"fuzzy_ok_{file_type}", items=len(sheet_items), conf=f"{result.confidence:.2f}")
                    return

        except Exception as e:
            result.errors.append(f"fuzzy_{file_type}: {e}")
            logger.warning(f"fuzzy_{file_type}_failed", error=str(e))

        fallback_method(file_path, result)  # _standard_xlsx_fallback или _standard_csv_fallback

    # парсинг с fuzzy

    @staticmethod
    def _parse_dataframe_with_fuzzy(df: pd.DataFrame) -> tuple[list[dict], dict]:
        """Парсинг файла с использованием Fuzzy для поиска данных."""

        if df.empty:
            return [], {}

        try:
            # ищем строку шапки: первая строка, где есть >=2 ячейки с буквами
            header_row: int = 0
            bonus_header: int = 0
            col_map: dict[str, int] = {}

            while True:
                # цикл для перебора строк шапки для поиска всех полей (multi-headers)

                while header_row < len(df) and df.iloc[header_row].astype(str).str.contains(r"[а-яa-z]", case=False,
                                                                                            regex=True,
                                                                                            na=False).sum() < 2:
                    header_row += 1

                headers = df.iloc[header_row].astype(str)

                COLUMN_KEYWORDS = {
                    "name": [" наименование ", " название ", "product", "item", " имя ", "название товара",
                             " номенклатура "],
                    "sku": ["артикул ", " арт.", "код ", "sku", "article", "тип ", "марка ", "марки ", "обозначение "],
                    "unit": ["ед. измерения", "единица измерения", "unit", "ед. изм"],
                    "quantity": ["количество", "кол-во", "кол.  "],
                    # unit price (generic)
                    "price_unit": ["цена ", "price", "расценк", "РРЦ ", "МРЦ "],
                    # explicit “без НДС”
                    "price_base": ["цена без ндс", "без ндс", "без НДС", "цена за ед без ндс"],
                    # explicit “с НДС”
                    "price_vat": ["цена с ндс", "с НДС", "вкл. НДС", "цена с учетом ндс"],
                    # totals
                    "total_no_vat": ["стоимость без ндс", "сумма без ндс", "итого без ндс"],
                    "total_vat": ["стоимость с ндс", "сумма с ндс", "итого с ндс"],
                    "total": ["стоимость", "сумма", "итого", "всего", "total", "amount"],
                    "manufacturer": ["производител", "бренд", "vendor", "manufacturer", "поставщик"],
                    "notes": ["примечани", "notes", "описание", "комментари", "характеристик"]
                }

                headers_list: list = headers.astype(str).tolist()

                def safe_lower_processor(value):
                    """Обработчик для приведения содержимого сроки к нижнему регистру без ошибок."""
                    return value.lower() if isinstance(value, str) else value

                for col_type, keywords in COLUMN_KEYWORDS.items():
                    # проходимся по каждому типу и ключевому слову

                    best_score = 0
                    best_idx = None

                    for keyword in keywords:
                        # проверяем совпадения по каждому слову и определяем лучшее
                        match = process.extractOne(keyword, headers_list, processor=safe_lower_processor,
                                                   scorer=fuzz.partial_ratio)

                        if match and match[1] > best_score:
                            best_score = match[1]
                            best_idx = match[2]
                    if best_idx is not None and best_score >= 80:
                        col_map[col_type] = best_idx

                if len(col_map) < 7 and bonus_header < 1:
                    # если не хватает значений и проверили только одну строку, проверяем следующую (multi-headers)
                    header_row += 1
                    bonus_header += 1
                    continue

                if not col_map or "name" not in col_map and "sku" not in col_map:
                    return [], {}

                break

            data = df.iloc[header_row + 1:].reset_index(drop=True)
            items: list[dict] = []

            for _, row in data.iterrows():
                # проходимся по каждой строке и парсим значения
                name_value = row.iloc[col_map["name"]] if "name" in col_map else ""
                sku_value = row.iloc[col_map["sku"]] if "sku" in col_map else ""
                unit_value = row.iloc[col_map["unit"]] if "unit" in col_map else ""
                manufacturer_value = row.iloc[col_map["manufacturer"]] if "manufacturer" in col_map else ""
                notes_value = row.iloc[col_map["notes"]] if "notes" in col_map else ""

                prices = collect_prices_for_row(row, col_map)

                quantity_value = row.iloc[col_map["quantity"]] if "quantity" in col_map else 0
                quantity = parse_price(quantity_value) or 0

                best_price = pick_best_price(prices, quantity or 0)

                item_values: dict = {
                    "name": str(name_value) if name_value and not pd.isna(name_value) else "",
                    "sku": str(sku_value) if sku_value and not pd.isna(sku_value) else "",
                    "unit": str(unit_value) if unit_value and not pd.isna(unit_value) else None,
                    "quantity": quantity,
                    "price": best_price,
                    "manufacturer": manufacturer_value,
                    "notes": notes_value if notes_value != name_value else ""
                }

                # создаем объект со всеми данными
                item = ExcelParser._build_item(item_values, raw_row=row, sku_col_idx=col_map.get("sku"))

                if item:
                    items.append(item)

            return items, col_map

        except IndexError:
            # если таблица маленькая и не получается найти нужный столбец, завершаем работу
            return [], {}

    @staticmethod
    def _handle_success(result: ParseResult, items: list, col_map: dict, method_suffix: str, sep: str = None):
        """Общий обработчик успеха fuzzy-парсинга."""
        result.items.extend(items)
        if not result.col_map and col_map:
            result.col_map = col_map
        result.method = f"{method_suffix}"
        result.confidence = quality(result.items)

    # fallback: стандартные парсеры

    def _standard_xlsx_fallback(self, file_path: Path, result: ParseResult):
        """Логика обработки XLSX-файлов с openpyxl, если Fuzzy не сработал."""

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                # проходимся по каждому листу в файле
                rows = list(wb[sheet_name].iter_rows(values_only=True))
                items = self._standard_parse(rows)

                if items:
                    result.items.extend(items)
                    result.method = "standard"

            wb.close()

        except Exception as e:
            result.errors.append(f"standard_xlsx_error: {e}")

        if result.items:
            result.confidence = quality(result.items)

    def _standard_xls_fallback(self, file_path: Path, result: ParseResult):
        """Логика обработки XLS-файлов с xlrd, если Fuzzy не сработал."""

        try:
            import xlrd
            wb = xlrd.open_workbook(str(file_path))

            for i in range(wb.nsheets):
                ws = wb.sheet_by_index(i)
                rows = [tuple(ws.cell_value(r, c) for c in range(ws.ncols))
                        for r in range(ws.nrows)]
                items = self._standard_parse(rows)

                if items:
                    # если объекты есть, записываем
                    result.items.extend(items)
                    result.method = "standard"

        except Exception as e:
            # если упало, записываем ошибку
            result.errors.append(f"standard_xls: {e}")

        if result.items:
            # если есть объекты, вычисляем качество обработки
            result.confidence = quality(result.items)

    def _standard_csv_fallback(self, file_path: Path, result: ParseResult):
        """Логика обработки CSV-файлов с pandas, если Fuzzy не сработал."""

        for sep in [";", ",", "\t", "|"]:
            # проходимся по каждому разделителю

            try:
                df = pd.read_csv(str(file_path), sep=sep, header=None, dtype=str)
                rows = [tuple(row) for _, row in df.iterrows()]
                items = self._standard_parse(rows)
                if items:
                    result.items.extend(items)
                    result.method = "csv_standard"
                    result.confidence = quality(items)
                    return
            except Exception as e:
                result.errors.append(f"csv_std({sep}): {e}")
                continue

    @staticmethod
    def _standard_parse(rows) -> list[dict] | None:
        """Стандартный парсинг."""

        # поиск заголовка
        for row_id, row in enumerate(rows):
            # проходимся по каждой строке

            col_map = simple_match_header(row)  # простым методом находим шапку

            # нужен хотя бы name или sku, плюс какой-то дополнительный столбец
            has_id = ("name" in col_map) or ("sku" in col_map)
            has_any_extra = any(k in col_map for k in ["unit", "quantity", "manufacturer", "notes"])

            if not has_id or not has_any_extra:
                # если нет наименования, артикля, и другого столбца, то идем дальше
                continue

            col_name = col_map.get("name")
            col_sku = col_map.get("sku")
            col_unit = col_map.get("unit")
            col_quantity = col_map.get("quantity")
            col_manufacturer = col_map.get("manufacturer")
            col_notes = col_map.get("notes")

            items: list[dict] = []

            for dr in rows[row_id + 1:]:

                # обозначаем переменные
                name_value = dr[col_name] if col_name is not None and col_name < len(dr) else ""
                sku_value = dr[col_sku] if col_sku is not None and col_sku < len(dr) else ""
                unit_value = dr[col_unit] if col_unit is not None and col_unit < len(dr) else ""
                manufacturer_value = dr[col_manufacturer] if col_manufacturer is not None and col_manufacturer < len(
                    dr) else ""
                notes_value = dr[col_notes] if col_notes is not None and col_notes < len(dr) else ""

                quantity_value = dr[col_quantity] if col_quantity is not None and col_quantity < len(dr) else 0
                quantity = parse_price(quantity_value) or 0

                prices: dict = collect_prices_for_row(dr, col_map)

                best_price = pick_best_price(prices, quantity or 0)

                item_values: dict = {
                    "name": str(name_value) if name_value and not pd.isna(name_value) else "",
                    "sku": str(sku_value) if sku_value and not pd.isna(sku_value) else "",
                    "unit": str(unit_value) if unit_value and not pd.isna(unit_value) else None,
                    "quantity": int(quantity_value) if quantity_value and not pd.isna(quantity_value) else 0,
                    "price": best_price,
                    "manufacturer": manufacturer_value if manufacturer_value and not pd.isna(manufacturer_value) else 0,
                    "notes": notes_value if notes_value != name_value else ""
                }

                item = ExcelParser._build_item(item_values)

                if item:
                    items.append(item)

            if items:
                return items

        return None

    @staticmethod
    def _build_item(item_values: dict, raw_row=None, sku_col_idx: Optional[int] = None) -> Optional[dict]:
        """
        Универсальная сборка item из сырых значений. name и sku могут быть пустыми; если оба пустые -> вернуть None.
        """

        name = item_values.get("name", "")
        sku = item_values.get("sku", "")

        def safe_extract(value):
            """Метод для корректной обработки NaN."""
            if pd.isna(value) or value == "":
                return ""
            if hasattr(value, 'item'):  # Series-like
                value = value.item()
            return str(value).strip()

        name = safe_extract(name)
        sku = safe_extract(sku)

        # базовая фильтрация
        if not name and not sku:
            return None
        if name and (len(name) < 3 or is_junk_row(name)):
            # если name есть, но мусорный — считаем, что позиции нет
            return None

        unit_value: str = item_values.get("unit")
        unit = normalize_unit(safe_extract(unit_value)) if unit_value else ""
        manufacturer = safe_extract(item_values.get("manufacturer", ""))
        notes = safe_extract(item_values.get("notes", ""))

        quantity = safe_extract(item_values.get("quantity", ""))

        # цена (с защитой от спутывания с SKU)
        price_value = item_values.get("price")
        parsed_price: float = parse_price(price_value) if price_value else None

        price: float = parsed_price or 0.0
        if parsed_price and raw_row is not None and sku_col_idx is not None:
            try:
                sku_raw = raw_row.iloc[sku_col_idx]
                sku_clean = safe_extract(sku_raw).replace("-", "").replace(" ", "")

                if sku_clean and str(int(parsed_price)) in sku_clean:
                    price = 0.0
                else:
                    price = parsed_price

            except (ValueError, IndexError):
                pass
        else:
            price = parsed_price or 0.0

        if not any([unit, price, manufacturer, notes]):
            return None

        item = {
            "name": name,
            "sku": sku,
            "unit": unit,
            "quantity": quantity,
            "price": price,
            "manufacturer": manufacturer,
            "notes": notes,
        }

        return item
